# -*- coding: utf-8 -*-
import os
import mysql.connector
import openpyxl
import uvicorn
import asyncio
from fastapi import FastAPI, HTTPException
from pydantic import BaseModel

from utils import compare_struct, is_file_exists, is_table_exists, get_fields, get_data, insert_into, update_operation_table, export_statistics_to_xlsx, get_column_comments, get_last_index, export_data_by_date_to_xlsx
from config import DB_CONFIG
from logger import logger
from schemas import Item, ExportTempData, ExportDataByDate
from fastapi.middleware.cors import CORSMiddleware

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"], 
    allow_headers=["*"],  
)


@app.post("/excel_process", responses={
    200: {
        'description': 'Успешно вставлено',
        'content': {'application/json': {'example': {'detail': 'Данные успешно вставлены в таблицу TABLE_NAME'}}}
    },
    400: {
        'description': 'Ошибка',
        'content': {'application/json': {'example': [
            {'detail': 'Ошибка при вставке данных в таблицу TABLE_NAME'},
            {'detail': 'Файл не найден'},
            {'detail': 'Таблица не найдена'},
            {'detail': 'Поля которые отсутсвуют или находятся под другим именем: '},
            {'detail': 'Последовательность полей не совпадает! Последовательность должна быть как в следующем: [поле1, поле2, поле3]'}
        ]}}
    }
})
async def process_excel_data(item: Item):
    cnx = mysql.connector.connect(**DB_CONFIG)
    cursor = cnx.cursor()

    file_path = item.file_path
    campaign_name = item.campaign_name
    call_date = item.call_date

    cursor.execute(
        'SELECT Input_Table, Result_Table FROM OPERATION WHERE Campaign_Name = %s', (campaign_name,))
    res = cursor.fetchone()

    if res:
        table_name = res[0]
        result_table_name = res[1]
    else:
        raise HTTPException(status_code=400, detail="Таблица не найдена")

    if not await is_file_exists(file_path=file_path):
        raise HTTPException(status_code=400, detail="Файл не найден")

    if not await is_table_exists(table_name=table_name):
        raise HTTPException(status_code=400, detail="Таблица не найдена")

    struct_comparison_result = await compare_struct(file_path=file_path, table_name=table_name)
    if not struct_comparison_result['status']:
        error_message = struct_comparison_result['message']
        missing_fields = struct_comparison_result.get('missing_fields', [])
        if missing_fields:
            error_message += f"Поля которые отсутсвуют или находятся под другим именем: {', '.join(missing_fields)}"
        raise HTTPException(status_code=400, detail=error_message)

    fields = await get_fields(sheet=table_name)
    data = await get_data(file_path=file_path, fields=fields)

    cursor.execute(
        'SELECT Campaign_Caption FROM OPERATION WHERE Campaign_Name = %s', (campaign_name,))
    app_name = cursor.fetchone()[0]

    query = f'DESCRIBE `{table_name}`;'
    cursor.execute(query)
    columns = cursor.fetchall()
    columns = [column[0] for column in columns if column[0] != 'index']
    columns = ', '.join(columns)

    if campaign_name == '3':
        query = f'SELECT Call_Date FROM `{table_name}`'
        cursor.execute(query)
        input_call_date = cursor.fetchone()
        if input_call_date is not None:
            input_call_date = input_call_date[0]
            query = f"DELETE FROM History_{table_name} WHERE Call_Date = '{input_call_date}'"
            cursor.execute(query)
            query = f"INSERT INTO History_{table_name} ({columns})  SELECT {columns} FROM {table_name} WHERE DATE(Call_Date) = '{input_call_date};"
            cursor.execute(query)

        query = f'TRUNCATE table {table_name}'
        cursor.execute(query)

    insert_result = await insert_into(
        cnx=cnx,
        cursor=cursor,
        table_name=table_name,
        data=data,
        app_name=app_name,
        fields=fields,
        campaign_name=campaign_name,
        call_date=call_date
    )

    cnx.commit()
    cursor.close()
    cnx.close()

    if not insert_result['status']:
        logger.error(insert_result['message'])
        raise HTTPException(
            status_code=400, detail='Ошибка при вставке данных.')

    return {"detail": f"Данные {insert_result['message']} контактов успешно загружены"}


@app.post("/crm_process", responses={
    200: {
        'description': 'Успешно вставлено',
        'content': {'application/json': {
            'example': {
                'detail': 'Данные успешно обработаны'
            }
        }
        }
    },
    400: {
        'description': 'Ошибка',
        'content': {'application/json': {'example': [
            {'detail': 'Ошибка: Описание ошибки'},
        ]}}
    }
})
async def process_crm_data():
    cnx = mysql.connector.connect(**DB_CONFIG)
    cursor = cnx.cursor()

    try:
        query = f"SELECT `Campaign_Caption`, `Result_Table`, `Input_Table` FROM `OPERATION` WHERE `Campaign_Name` = '4'"
        cursor.execute(query)
        res = cursor.fetchone()

        app_name = res[0]
        res_table_name = res[1]
        input_table_name = res[2]

        query = f'DESCRIBE `COLLECTION`;'
        cursor.execute(query)
        columns = cursor.fetchall()
        columns = [
            f'`{column[0]}`' for column in columns if column[0] != 'index']
        columns_string = ', '.join(columns)

        logger.info('Deleting old data from result table')
        res_delete_query = f"TRUNCATE TABLE {res_table_name}"
        cursor.execute(res_delete_query)


        last_date_query = f"SELECT Date_Upload FROM `{input_table_name}`"
        cursor.execute(last_date_query)
        last_date_query_result = cursor.fetchone()
        if last_date_query_result is not None:
            logger.info('Replacing history input data with input data')
            last_date = last_date_query_result[0]
            delete_history_query = f"DELETE FROM `History_{input_table_name}` WHERE Date_Upload = '{last_date}'"
            cursor.execute(delete_history_query)

            last_index = await get_last_index(f'History_{input_table_name}')
            last_index += 1
            row_insert_query = f'SET @row_number = {last_index};'
            cursor.execute(row_insert_query)
            insert_query = f'''
                                INSERT INTO `History_{input_table_name}` 
                                (`index`, {columns_string}) 
                                SELECT
                                    (@row_number := @row_number + 1) AS `index`, 
                                    {columns_string} 
                                FROM `{input_table_name}`;
                            '''
            cursor.execute(insert_query)
            cnx.commit()
            logger.info('History input data replaced with input data')
        
        logger.info('Deleting old data from input table')
        delete_query = f"TRUNCATE TABLE {input_table_name}"
        cursor.execute(delete_query)

        row_insert_query = 'SET @row_number = 0;'
        cursor.execute(row_insert_query)
        insert_query = f'''
                    INSERT INTO `{input_table_name}` 
                    (`index`, {columns_string}) 
                    SELECT 
                        (@row_number := @row_number + 1) AS `index`,
                        {columns_string} 
                    FROM 
                        `COLLECTION` 
                    ORDER BY 
                        `Priority` ASC, `Debt` DESC, `Call_Start_Time`;
                        '''
        cursor.execute(insert_query)
        cnx.commit()

        last_index = await get_last_index(f'History_{input_table_name}')
        last_index += 1
        row_insert_query = f'SET @row_number = {last_index};'
        cursor.execute(row_insert_query)
        insert_query = f'''
                            INSERT INTO `History_{input_table_name}` 
                            (`index`, {columns_string}) 
                            SELECT
                                (@row_number := @row_number + 1) AS `index`, 
                                {columns_string} 
                            FROM `COLLECTION` 
                            ORDER BY `Priority` ASC, `Debt` DESC, `Call_Start_Time`
                        '''
        cursor.execute(insert_query)
        cnx.commit()

        update_app_name_query = f"UPDATE {input_table_name} SET App_Name = '{app_name}'"
        cursor.execute(update_app_name_query)
        cnx.commit()

        res_of_update_operation = await update_operation_table(table_name=input_table_name)

        if res_of_update_operation:
            message = "CRM-данные обработаны."
            logger.info('CRM data distributed successfully!')
        else:
            message = "Ошибка при обработке CRM-данных."
            logger.error('Error while distributing CRM data!')
            logger.error(message)
            raise HTTPException(
                status_code=500, detail=f"Ошибка")
        return {"detail": f'{message}'}
    except Exception as e:
        logger.exception(e)
        raise HTTPException(status_code=500, detail=f"Ошибка")
    finally:
        cnx.commit()
        cursor.close()
        cnx.close()


@app.post('/export_temp_data')
async def export_temp_data(item: ExportTempData):
    cnx = mysql.connector.connect(**DB_CONFIG)
    cursor = cnx.cursor()
    campaign_name = item.campaign_name
    try:

        result_file_path, input_file_path = await export_statistics_to_xlsx(campaign_name=campaign_name)
        return {
            'status': True,
            'result_file_path': result_file_path,
            'input_file_path': input_file_path,
        }

    except Exception as e:
        logger.error(e)
        raise HTTPException(status_code=500, detail=f"Ошибка: {str(e)}")
    finally:
        cnx.commit()
        cursor.close()
        cnx.close()


@app.post('/export_data_by_date')
async def export_data_by_date(item: ExportDataByDate):
    cnx = mysql.connector.connect(**DB_CONFIG)
    cursor = cnx.cursor()
    campaign_name = item.campaign_name
    start_date = item.start_date
    end_date = item.end_date
    try:
        result_file_path, input_file_path = await export_data_by_date_to_xlsx(campaign_name=campaign_name, start_date=start_date, end_date=end_date)
        return {
            'status': True,
            'result_file_path': result_file_path,
            'input_file_path': input_file_path,
        }

    except Exception as e:
        logger.error(e)
        raise HTTPException(status_code=500, detail=f"Ошибка: {str(e)}")
    finally:
        cnx.commit()
        cursor.close()
        cnx.close()


@app.post('/collection_process')
async def process_collection(item: Item):
    try:
        file_path = item.file_path
        table_name = 'COLLECTION'
        cnx = mysql.connector.connect(**DB_CONFIG)

        cursor = cnx.cursor()

        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        columns = []
        for i in range(1, sheet.max_column + 1):
            columns.append(sheet.cell(1, i).value)

        for i in range(len(columns)):
            columns[i] = f'`{columns[i]}`'

        count = 1
        query = f"Truncate TABLE {table_name};"
        cursor.execute(query)
        cnx.commit()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue

            query = "INSERT INTO {0} ({1}) VALUES ({2});".format(
                table_name, ', '.join(columns), ', '.join(['%s'] * len(columns)))
            row = list(row)
            row = tuple(row)
            cursor.execute(query, row)
            cnx.commit()
            count += 1
        logger.info('SUCCESFULLY INSERTED TO COLLECTION')
        cnx.commit()
        cursor.close()
        cnx.close()
        return {"detail": "Данные успешно загружены в таблицу COLLECTION"}
    except Exception as e:
        logger.error(e)
        raise HTTPException(status_code=500, detail=f"Ошибка: {str(e)}")


if __name__ == "__main__":
    uvicorn.run('main:app', host="localhost", port=8000, reload=True)
